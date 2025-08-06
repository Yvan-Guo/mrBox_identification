# -*- coding: utf-8 -*-
"""
Created on Mon Sep 16 23:32:58 2024

@author: Yvan_Guo
"""

import os
import pandas as pd
import numpy as np
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor
from Bio import SeqIO
import subprocess
from scipy.stats import spearmanr, wilcoxon, fisher_exact
import yaml
import shutil

def preprocess_data(config):
    """
    Preprocesses DNA methylation data and annotations for downstream analysis.
    
    Loads CpG annotations, DNA sequences, and methylation data from specified files,
    merges them, and filters out invalid entries. This step ensures data integrity
    for subsequent correlation and BLAST analyses.
    
    Args:
        config (dict): Configuration dictionary with file paths and settings.
    
    Returns:
        tuple: (cpg_id, seqset, met_ds) containing annotation DataFrame, sequence dictionary,
        and merged methylation dataset.
    """
    workdir = config.get('workdir', '/workdir')
    os.chdir(workdir)
    
    try:
        # Load and preprocess annotations
        cpg_url = config.get('cpg_annotation_file', 'illuminaMethyl450_hg38_GDC.txt')
        cpg_id = pd.read_csv(cpg_url, sep='\t', comment='#', names=["Gene", "Chrosome", "Site"])
        cpg_id = cpg_id[cpg_id["Chrosome"] != '*']
        
        # Load and preprocess DNA sequences
        seq_file = config.get('sequence_file', 'CpG_sequence_set.txt')
        seqset = SeqIO.to_dict(SeqIO.parse(seq_file, 'fasta'))
        anno = pd.DataFrame({"names": list(seqset.keys())}).drop_duplicates().set_index('names')
        
        # Load and preprocess methylation data
        met_file = config.get('methylation_file', 'PANCAN_HumanMethylation450.tsv.gz')
        met_ds = pd.read_csv(met_file, sep='\t', compression='gzip', index_col=0)
        
        # Merge annotations with methylation data
        met_ds = met_ds.merge(anno, left_index=True, right_index=True)
        met_ds.drop(columns=["names"], inplace=True)
        
        return cpg_id, seqset, met_ds
    except Exception as e:
        print(f"Error in data preprocessing: {str(e)}")

def cor_frame(r, met_ds, pth1, grid_size=1250, cor_type='spearman'):
    """
    Computes correlation matrix for a subset of methylation data using a grid-based approach.
    
    Args:
        r (int): Row index for grid processing.
        met_ds (pd.DataFrame): Methylation dataset.
        pth1 (str): Path to save correlation results.
        grid_size (int): Size of data grid for memory-efficient processing. Default is 1250,
            chosen based on memory constraints and dataset size (see manuscript for details).
        cor_type (str): Type of correlation ('spearman' or 'pearson'). Default is 'spearman'
            due to non-normal distribution of methylation data.
    """
    a0 = (r - 1) * grid_size
    a1 = min(a0 + grid_size, met_ds.shape[0])
    a = range(a0, a1)
    cor = np.zeros((met_ds.shape[0], len(a)))
    p_cor = np.zeros_like(cor)
    
    for k in range((met_ds.shape[0] + grid_size - 1) // grid_size):
        b0 = k * grid_size
        b1 = min((k + 1) * grid_size, met_ds.shape[0])
        b = range(b0, b1)
        
        if len(a) == len(b):
            test = met_ds.iloc[np.r_[a, b],:].T
            r_values, p_values = spearmanr(test)
            raw_p_values = p_values[len(a):, :len(a)]
            adjusted_p_values = np.minimum(1., 0.05 * len(raw_p_values) / (np.arange(len(raw_p_values)) + 1))
            p_cor[b, :] = np.where(adjusted_p_values <= 0.05, 1, 0)
    
    cor = np.multiply(cor, p_cor)
    np.savetxt(f'{pth1}/{r}.csv', cor, delimiter=",")

def blast_analysis(input_file, db, output_file, config):
    """
    Performs BLAST analysis for sequence similarity search.
    
    Args:
        input_file (str): Path to input query file.
        db (str): Path to BLAST database.
        output_file (str): Path to save BLAST results.
        config (dict): Configuration dictionary with BLAST settings.
    """
    blastn = config.get('blastn_path', 'blastn.exe')
    num_threads = config.get('blast_threads', 32)
    try:
        subprocess.run([blastn, 
                        "-db", db, 
                        "-query", input_file, 
                        "-out", output_file, 
                        "-outfmt", "7", 
                        "-num_threads", str(num_threads)])
    except Exception as e:
        print(f"Error in BLAST analysis: {str(e)}")
        raise

def execute_blast_analysis(loc, pth, pth1):
    """
    Executes BLAST analysis for sequence similarity using configured paths.
    
    Args:
        loc (str): Location identifier.
        pth (str): Path for temporary files.
        pth1 (str): Path for output files.
    """
    def finalseq_chr_py(loc, pth1):
        from openpyxl import load_workbook
        seq = ''
        target_file = os.path.join('temp', str(loc), 'Target.txt')
        try:
            with open(target_file, 'r') as f:
                tar = f.readline()[:-1]
                for line in f.readlines():
                    seq += line.strip('\n')
            wb = load_workbook(os.path.join('temp', str(loc), 't_blast_temp.xlsx'))
            ws = wb[wb.sheetnames[0]]
            rm = ws.max_row
            st, en = [], []
            for r in range(2, rm+1):
                st.append(ws.cell(r, 1).value)
                en.append(ws.cell(r, 2).value)
            if not st or not en:
                return 1
            with open(os.path.join(pth1, 'Significant_sequence_cluster_short.txt'), 'a') as f:
                for k in range(len(st)):
                    fr, to = st[k], en[k]+1
                    sequence = seq[fr:to]
                    f.write(f"{tar}_{fr}_{to}\n")
                    f.write(f"{sequence}\n\n")
            return 0
        except Exception as e:
            print(f"Error in finalseq_chr_py: {str(e)}")
            return 1
    
    def parse_blast_output(output_file):
        try:
            return pd.read_csv(output_file, sep='\t', comment='#', header=None)
        except FileNotFoundError:
            return pd.DataFrame()
    
    posb_file = os.path.join(pth, 'PosB_result.txt')
    negb_file = os.path.join(pth, 'NegB_result.txt')
    blast_analysis(os.path.join(pth, 'result.txt'), os.path.join(pth, 'Pos'), posb_file)
    posb_output = parse_blast_output(posb_file)
    posb_counts = posb_output[1].value_counts()
    blast_analysis(os.path.join(pth, 'result.txt'), os.path.join(pth, 'Neg'), negb_file)
    negb_output = parse_blast_output(negb_file)
    negb_counts = negb_output[1].value_counts()
    results = pd.DataFrame({
        'Row': posb_counts.index,
        'Pos': posb_counts.values,
        'Neg': negb_counts.reindex_like(posb_counts).fillna(0).values
    })
    filtered_results = results[results['Pos'] > 25]
    filtered_results = filtered_results.apply(lambda x: {
        **x, "st": x['Row'] // 1e7, "en": x['Row'] % 1e7
    }, axis=1)
    final_results = filtered_results[filtered_results.apply(lambda x: fisher_exact([[x['Pos'], x['Neg']], [100-x['Pos'], 100-x['Neg']]])[1] <= 0.05, axis=1)]
    
    if not final_results.empty:
        final_results[['st', 'en']].to_excel(os.path.join(pth, "t_blast_temp.xlsx"))
        finalseq_chr_py(loc, pth1)

def predict_mrdna(seqset, sq_r, col_n, config):
    """
    Predicts methylation-related DNA sequences using slicing and BLAST analysis.
    
    Args:
        seqset (dict): Dictionary of DNA sequences.
        sq_r (pd.DataFrame): DataFrame with correlation results.
        col_n (int): Column index for processing.
        config (dict): Configuration dictionary.
    """
    def slice_py(loc, pth):
        sequence = ''
        target_file = os.path.join(pth, 'Target.txt')
        try:
            with open(target_file, 'r') as f:
                f.readline()
                for line in f.readlines():
                    sequence += line.strip('\n')
            le, step, fr, to = 200, 100, 1, 5000
            pre_seq = sequence[fr:to]
            with open(os.path.join(pth, f'Slice_{fr}_{to}.txt'), 'w') as f:
                for r in range(0, len(pre_seq) - le + 1, step):
                    fragment = pre_seq[r:r + le]
                    f.write(f'>{r + fr}\n')
                    f.write(f'{fragment}\n')
        except Exception as e:
            print(f"Error in slice_py: {str(e)}")
            raise
    
    def extract_py(loc, pth):
        from openpyxl import load_workbook
        seq = ''
        target_file = os.path.join(pth, 'Target.txt')
        try:
            with open(target_file, 'r') as f:
                tar = f.readline()[:-1]
                for line in f.readlines():
                    seq += line.strip('\n')
            wb = load_workbook(os.path.join(pth, 'extracted_temp.xlsx'))
            ws = wb[wb.sheetnames[0]]
            rm = ws.max_row
            st, en = [], []
            for r in range(2, rm+1):
                lo = ws.cell(r, 1).value
                st.append(lo)
                en.append(lo + 200)
            if not st or not en:
                return 1
            if en[-1] > 8000:
                en[-1] = 8000
            if st[0] < 1:
                st[0] = 1
            with open(os.path.join(pth, 'result.txt'), 'w') as f:
                for k in range(len(st)):
                    fr, to = st[k], en[k]+1
                    sequence = seq[fr:to]
                    f.write(f'>{fr*10000000 + to}\n')
                    f.write(f'{sequence}\n')
            with open(os.path.join(pth, 'G_sequence_cluster.txt'), 'a') as f:
                for k in range(len(st)):
                    fr, to = st[k], en[k]+1
                    sequence = seq[fr:to]
                    f.write(f'{tar}_{fr*10000000 + to}\n')
                    f.write(f'{sequence}\n\n')
            return 0
        except Exception as e:
            print(f"Error in extract_py: {str(e)}")
            return 1
    
    loc = sq_r.columns[col_n]
    pth = f"temp/{loc}"
    os.makedirs(pth, exist_ok=True)
    
    cor = sq_r[['Site', sq_r.columns[col_n]]].sort_values(by=['Rvalue'], ascending=False).dropna()
    top_sites = cor.iloc[1:101, 0].values
    bottom_sites = cor.iloc[-100:, 0].values
    
    fasta_files = {
        "Target.txt": seqset[loc],
        "Pos.txt": [seqset[site] for site in top_sites],
        "Neg.txt": [seqset[site] for site in bottom_sites]
    }
    for filename, records in fasta_files.items():
        SeqIO.write(records, os.path.join(pth, filename), "fasta")
    
    slice_py(loc, pth)
    makeblastdb = config.get('makeblastdb_path', 'makeblastdb.exe')
    blastn = config.get('blastn_path', 'blastn.exe')
    
    for filename in ["Pos.txt", "Neg.txt"]:
        subprocess.run([makeblastdb, "-in", os.path.join(pth, filename), "-dbtype", "nucl", "-out", os.path.join(pth, filename)])
    
    subprocess.run([blastn, "-db", os.path.join(pth, "Pos"), "-query", os.path.join(pth, "Slice_1_8000.txt"), "-out", os.path.join(pth, "Slice_P.txt"), "-outfmt", "7", "-num_threads", "2"])
    
    try:
        output = pd.read_csv(os.path.join(pth, 'Slice_P.txt'), sep='\t', comment='#', header=None)
    except FileNotFoundError:
        output = None
    
    if output is not None:
        p_slice = output.groupby(0).filter(lambda x: x.shape[0] > 10).assign(Dataset="Pos.")
        if not p_slice.empty:
            wil_p = p_slice.groupby([1]).apply(lambda x: wilcoxon(x[11], alternative='greater').pvalue).reset_index().rename(columns={0: 'Loc_l', 1: 'P'})
            wil_p['Signif'] = wil_p['P'].apply(lambda x: 0 if pd.isna(x) else x)
            significant_p = wil_p[wil_p['Signif'] <= 0.05]
            if not significant_p.empty:
                significant_p.to_excel(os.path.join(pth, 'extracted_temp.xlsx'))
                result = extract_py(loc, pth)
                if result != 1:
                    execute_blast_analysis(loc, pth, pth)
    shutil.rmtree(pth)

def main(config_file):
    """
    Main execution function to orchestrate data preprocessing, correlation analysis,
    and sequence prediction based on configuration settings.
    
    Args:
        config_file (str): Path to YAML configuration file with parameters and paths.
    """
    try:
        with open(config_file, 'r') as file:
            config = yaml.safe_load(file)
        
        cpg_id, seqset, met_ds = preprocess_data(config)
        grid_size = config.get('grid_size', 1250)
        num_chr = dict(cpg_id["Chrosome"].value_counts())
        
        for chr_idx, chr_name in enumerate(num_chr.keys()):
            temp_anno = cpg_id[cpg_id["Chrosome"] == chr_name]
            met = met_ds.loc[temp_anno.index]
            pth1 = os.path.join('cor_data_row', f'cor_frame_{chr_name}')
            os.makedirs(pth1, exist_ok=True)
            l_met = met.shape[0]
            loop = (l_met + grid_size - 1) // grid_size
            
            if len([f for f in os.listdir(pth1) if f.endswith(".csv")]) < loop:
                print("Building correlation database...")
                max_workers = config.get('max_cores', 25)
                with ProcessPoolExecutor(max_workers=max_workers) as executor:
                    futures = [executor.submit(cor_frame, r, met, pth1, grid_size) for r in range(1, loop + 1)]
                    for future in futures:
                        future.result()
                if len([f for f in os.listdir(pth1) if f.endswith(".csv")]) < loop:
                    print("Cannot build correlation database, please try fewer cores!")
            
            for n_cor in range(1, loop + 1):
                print("Enrichment analysis...")
                sq_r = pd.read_csv(os.path.join(pth1, f'{n_cor}.csv'))
                max_thread_workers = config.get('max_thread_workers', 45)
                with ThreadPoolExecutor(max_workers=max_thread_workers) as executor:
                    futures = [executor.submit(predict_mrdna, seqset, sq_r, col_n, config) for col_n in range(2, sq_r.shape[1])]
                    for future in futures:
                        future.result()
    except Exception as e:
        print(f"Error in main execution: {str(e)}")

if __name__ == "__main__":
    config_path = "config.yaml"
    main(config_path)
